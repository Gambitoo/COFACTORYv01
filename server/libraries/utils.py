import numpy as np
from io import BytesIO
from datetime import datetime, timedelta, time
import time as tm
import pyodbc
from openpyxl import Workbook, load_workbook

class TimeUnit:
    id = 0
    GR_instances, PT_instances = [], []
    def __init__(self, Machine):
        TimeUnit.id += 1
        self.id = TimeUnit.id
        self.Name = f"timeUnit_{TimeUnit.id}"
        self.Machine = Machine
        self.ExecutionPlans = [] #Not inserted into the DB
        self.ST = 0
        self.CoT = 0

    @classmethod
    def clear_instances(cls):
        # Clears all instances from both lists
        cls.PT_instances.clear()  
        cls.GR_instances.clear()  
        cls.id = 0  # Optionally reset the id counter if needed
    
    """@classmethod
    def clear_new_instances(cls):
        cls.new_instances.clear()
        cls.id = 0"""
        
    def sort_by_need(self, torc_solution):
        """
        Calculate priority based on when TORC items need this TREF item.
        Returns timestamp of earliest TORC operation that needs any item from this TimeUnit.
        Lower values = higher priority (needed sooner)
        """
        if not torc_solution:
            return 0

        urgency_score = 0
        earliest_need = datetime.min
        
        for _, operations in torc_solution.items():
            for op in operations:
                _, details = op
                _, exec_plan, start_time, _ = details

                # Check if this time unit produces items needed by this TORC operation
                for tu_exec_plan in self.ExecutionPlans:
                    if (tu_exec_plan.ItemRoot and 
                        tu_exec_plan.ItemRoot.Name == exec_plan.ItemRelated.Name and
                        tu_exec_plan.ProductionOrder.id == exec_plan.ProductionOrder.id):

                        earliest_need = min(earliest_need, start_time)

        if earliest_need:
            # Convert to urgency score (earlier = higher score)
            urgency_score = (earliest_need - datetime.min).total_seconds()

        return urgency_score

    def calculate_average_due_date(self):
        """Organize the Products by Due Date"""
        due_dates = [exec_plan.ProductionOrder.DD for exec_plan in self.ExecutionPlans]
        if due_dates:
            seconds = [tm.mktime(d.timetuple()) for d in due_dates]
            avg_due_date = datetime.fromtimestamp(np.mean(seconds))
            return avg_due_date
        else:
            return None

    def get_average_diameter(self, database):
        """Get the average diameter of the machine"""
        diameters = [int(Items.get_Item(ep.ItemRelated.Name, database).Diameter * 1000) for ep in self.ExecutionPlans if
                     Items.get_Item(ep.ItemRelated.Name, database)]
        if diameters:
            return np.mean(diameters)
        else:
            return float('inf')

    def get_primary_material_type(self):
        """Determine the primary material type among ExecutionPlans."""
        material_types = {ep.ItemRelated.MaterialType for ep in self.ExecutionPlans}
        if material_types:
            return sorted(material_types)[0]  # Assuming a sorted order can determine primary type
        return None

    def get_average_weight(self):
        """Get the average weight for execution plans with the same due date."""
        weights = [exec_plan.ProductionOrder.Weight for exec_plan in self.ExecutionPlans]
        if weights:
            return np.mean(weights)
        return float('inf')

    def calculate_time(self, TU_ST, previous_TU, data_handler, current_time):
        shift_start_times = [time(0, 0), time(8, 0), time(16, 0)]  # Midnight, 8 AM, 4 PM

        def start_time():
            with pyodbc.connect(data_handler.ConnectionString) as conn:
                with conn.cursor() as cursor:  # Opens the cursor
                    # Check if TimeUnits table has records for this machine
                    cursor.execute("SELECT TOP 1 id, CompletionTime FROM TimeUnits WHERE Machine = ? ORDER BY CompletionTime DESC", (self.Machine,))
                    result = cursor.fetchone()
                    TU_id, latest_ep_CoT = result if result else (None, None)
    
                    if TU_id:
                        cursor.execute("SELECT ExecutionPlanId FROM TimeUnitExecutionPlans WHERE TimeUnitId = ?", (TU_id,))
                        exec_plan_id = cursor.fetchone()[0]
    
                        cursor.execute("SELECT Item FROM ExecutionPlans WHERE id = ?", (exec_plan_id,))
                        previous_plan_item = cursor.fetchone()[0]
    
                        previous_type = next((item.MaterialType for item in data_handler.Items if item.Name == previous_plan_item), None)
                    else:
                        previous_type = None

            # Determine the initial start time
            start_time = max(latest_ep_CoT, current_time) if latest_ep_CoT else current_time

            # Check if a setup time is needed
            current_type = self.ExecutionPlans[0].ItemRelated.MaterialType
            if current_type != previous_type:
                setup_time = next(
                    (float(instance.SetupTime) for instance in data_handler.SetupTimesByMaterial
                     if instance.FromMaterial == previous_type and instance.ToMaterial == current_type), 0.0)
                start_time += timedelta(hours=setup_time)

            return next_shift_start_time(start_time)

        def next_shift_start_time(current_time):
            """Calculate next available shift start time"""
            current_date = current_time.date()
        
            for shift_start in shift_start_times:
                shift_start_dt = datetime.combine(current_date, shift_start)
                if current_time < shift_start_dt:
                    return shift_start_dt
            return datetime.combine(current_date + timedelta(days=1), shift_start_times[0])

        # Determine start time (ST)
        if previous_TU:
            setup_time = None
            previous_type = previous_TU.ExecutionPlans[0].ItemRelated.MaterialType
            current_type = self.ExecutionPlans[0].ItemRelated.MaterialType
            if previous_type != current_type:
                setup_time = next(
                    (float(instance.SetupTime) for instance in data_handler.SetupTimesByMaterial
                     if instance.FromMaterial == previous_type and instance.ToMaterial == current_type), 0.0)
            self.ST = previous_TU.CoT + timedelta(hours=setup_time) if setup_time else previous_TU.CoT
        else:
            if TU_ST is None:
                self.ST = start_time()
            else:
                calculated_ST = start_time()
                if calculated_ST > TU_ST:
                    self.ST = calculated_ST
                else:
                    self.ST = TU_ST

        # Get max completion time (CoT)      
        max_CT = max(
            (next((routing.CycleTime / 1000) for routing in data_handler.Routings
                  if routing.Item == exec_plan.ItemRelated.Name and routing.Machine == self.Machine)
             * exec_plan.Quantity) for exec_plan in self.ExecutionPlans
        )
        self.ST += timedelta(minutes=(max_CT * 0.16))
        self.CoT = self.ST + timedelta(minutes=max_CT)

class BoM:
    id = 0
    GR_instances, PT_instances = [], []
    def __init__(self, ItemRoot, BoMQuantity, BoMQuantityUnit, Revision, Default):
        BoM.id += 1
        self.id = BoM.id
        self.ItemRoot = ItemRoot
        self.BoMQuantity = BoMQuantity
        self.BoMQuantityUnit = BoMQuantityUnit
        self.BoMItems = []
        self.Revision = Revision
        self.Default = Default

    @classmethod
    def clear_instances(cls):
        # Clears all instances from both lists
        cls.PT_instances.clear()  
        cls.GR_instances.clear()  
        cls.id = 0  # Optionally reset the id counter if needed

    def add_BoM_items(self, *args):
        self.BoMItems.extend(args)

class BoMItem:
    id = 0
    GR_instances, PT_instances = [], []
    def __init__(self, ItemRoot, ItemRelated, Quantity, NetQuantity, NetQuantityUnit, Database):
        BoMItem.id += 1
        self.id = BoMItem.id
        self.ItemRoot = ItemRoot
        self.ItemRelated = ItemRelated
        self.NetQuantity = NetQuantity
        self.NetQuantityUnit = NetQuantityUnit
        self.Quantity = Quantity
        
        if Database == "COFACTORY_GR":
            BoMItem.GR_instances.append(self)
        elif Database == "COFACTORY_PT":
            BoMItem.PT_instances.append(self)
        
    @classmethod
    def clear_instances(cls):
        # Clears all instances from both lists
        cls.PT_instances.clear()  
        cls.GR_instances.clear()  
        cls.id = 0  # Optionally reset the id counter if needed

class ProductionOrder:
    id = 0
    def __init__(self, Product, Quantity, DD, Weight):
        ProductionOrder.id += 1
        self.id = ProductionOrder.id
        self.Name = f"Operation_{ProductionOrder.id}"
        self.Product = Product
        self.Quantity = Quantity
        self.DD = DD
        self.ST = 0
        self.CoT = 0
        self.Weight = Weight #This field will not be placed in the DB

class ExecutionPlan:
    id = 0
    GR_instances, PT_instances = [], []
    def __init__(self, ItemRoot, ItemRelated, Quantity, BoMId, ProductionOrder):
        ExecutionPlan.id += 1
        self.id = ExecutionPlan.id
        self.ItemRoot = ItemRoot
        self.ItemRelated = ItemRelated
        self.Quantity = Quantity
        self.Machine = None
        self.Position = None
        self.ProductionOrder = ProductionOrder
        self.ST = 0
        self.CoT = 0
        self.BoMId = BoMId #Not inserted into the DB
        self.PlanoId = None
        
    @classmethod
    def clear_instances(cls):
        # Clears all instances from both lists
        cls.PT_instances.clear()  
        cls.GR_instances.clear()  
        cls.id = 0  # Optionally reset the id counter if needed

    """@classmethod
    def clear_new_instances(cls):
        cls.new_instances.clear()  # Clears all instances from the list
        cls.id = 0  # Optionally reset the id counter if needed"""

    """@classmethod
    def remove_by_id(cls, target_id):
        cls.new_instances = [instance for instance in cls.new_instances if instance.id != target_id]
        
    @classmethod
    def remove_by_bomId(cls, target_bomId):
        cls.new_instances = [instance for instance in cls.new_instances if instance.BoMId != target_bomId]"""

class Machines:
    GR_instances, PT_instances = [], []
    def __init__(self, MachineCode, Input, Output, RunningTimeFactor):
        self.MachineCode = MachineCode
        self.Input = Input
        self.Output = Output
        self.RunningTimeFactor = RunningTimeFactor
        self.IsActive = True
        
    @classmethod
    def clear_instances(cls):
        # Clears all instances from both lists
        cls.PT_instances.clear()  
        cls.GR_instances.clear()  
        cls.id = 0  # Optionally reset the id counter if needed

class Routings:
    GR_instances, PT_instances = [], []
    def __init__(self, Item, Machine, CycleTime, Weight):
        self.Item = Item
        self.Machine = Machine
        self.CycleTime = CycleTime
        self.Weight = Weight

class Items:
    id = 0
    GR_instances, PT_instances = [], []
    def __init__(self, Name, MaterialType, Unit, Input, Diameter, Process, OrderIncrement):
        Items.id += 1
        self.ID = Items.id
        self.Name = Name
        self.MaterialType = MaterialType
        self.Unit = Unit
        self.Process = Process
        self.Input = Input if Input is not None else 0
        self.Diameter = Diameter
        self.Category = None
        self.OrderIncrement = OrderIncrement
    
    @classmethod
    def get_Item(cls, name, database):
        instances = cls.GR_instances if database == "COFACTORY_GR" else cls.PT_instances
        return next((instance for instance in instances if instance.Name == name), None)

class SetupTimesByMaterial:
    GR_instances, PT_instances = [], []
    def __init__(self, FromMaterial, ToMaterial, SetupTime):
        self.FromMaterial = FromMaterial
        self.ToMaterial = ToMaterial
        self.SetupTime = SetupTime

class Stock:
    GR_instances, PT_instances = [], []
    def __init__(self, Warehouse, Item, StockAvailable, StockAllocated, StockEconomic):
        self.Warehouse = Warehouse
        self.Item = Item
        self.StockAvailable = StockAvailable
        self.StockAllocated = StockAllocated
        self.StockEconomic = StockEconomic
        
class LN_ProductionOrders:
    GR_instances, PT_instances = [], []
    def __init__(self, ID, Item, Routing, Quantity, ST, CoT):
        self.ID = ID
        self.Item = Item
        self.Routing = Routing
        self.Quantity = Quantity
        self.ST = ST
        self.CoT = CoT

class DataHandler:
    def __init__(self, database, connection_string):
        self.Database, self.ConnectionString, self.CurrentTime = database, connection_string, None
        # Store user specific and general data
        self.ExecutionPlans, self.TimeUnits, self.ProductionOrders = [], [], []
        self.Machines = Machines.GR_instances if self.Database == "COFACTORY_GR" else Machines.PT_instances
        self.BoMs = BoM.GR_instances if self.Database == "COFACTORY_GR" else BoM.PT_instances
        self.BoMItems = BoMItem.GR_instances if self.Database == "COFACTORY_GR" else BoMItem.PT_instances
        self.Routings = Routings.GR_instances if self.Database == "COFACTORY_GR" else Routings.PT_instances
        self.Items = Items.GR_instances if self.Database == "COFACTORY_GR" else Items.PT_instances
        self.SetupTimesByMaterial = SetupTimesByMaterial.GR_instances if self.Database == "COFACTORY_GR" else SetupTimesByMaterial.PT_instances
        self.Stock = Stock.GR_instances if self.Database == "COFACTORY_GR" else Stock.PT_instances
        # Store process specific data
        self.RODMachines, self.TorcMachines, self.TrefMachines = [], [], []
        self.RODItems, self.TorcItems, self.TrefItems = [], [], []
        self.RODSolution, self.TorcSolution = None, None
        self.Criteria = {}
    
    def removeEPbyID(self, target_id):
        self.ExecutionPlans = [instance for instance in self.ExecutionPlans if instance.id != target_id]
        
    def removeEPbyBoMID(self, target_bomId):
        self.ExecutionPlans = [instance for instance in self.ExecutionPlans if instance.BoMId != target_bomId]

    def getInputBoMs(self, file_name):
        item_root_boms = {}
        Extrusion_Input = load_workbook(file_name)
        Extrusion_Input_Active = Extrusion_Input.active
        unique_product_names = set()
        for row in Extrusion_Input_Active.iter_rows(min_row=2, values_only=True):
            product_name = row[0]
            item_root = Items.get_Item(product_name, self.Database)
            if item_root:
                if product_name not in unique_product_names:
                    unique_product_names.add(product_name)
                for bom in self.BoMs:
                    if bom.ItemRoot == item_root.Name:
                        for BoM_Item in bom.BoMItems:
                            item_related = Items.get_Item(BoM_Item.ItemRelated, self.Database)
                            if item_related and item_related.Process == "BUN" and item_related.Name not in unique_product_names:
                                unique_product_names.add(item_related.Name)
        for item_root in unique_product_names:
            match_count = 0  # Track the number of BoMs for this root Item
            item_root_boms[item_root] = []
            for bom in self.BoMs:
                if bom.ItemRoot == item_root:
                    match_count += 1
                    bom_items = []
                    # Add each BoM to the bom_items list
                    for BoM_Item in bom.BoMItems:
                        item_related = Items.get_Item(BoM_Item.ItemRelated, self.Database)
                        for _ in range(BoM_Item.Quantity):
                            bom_items.append(item_related.Name)
                    item_root_boms[item_root].append(bom_items)
            # If match_count <= 1, then it means root Item only has one BoM, so it is removed from item_root_boms
            if match_count <= 1:
                item_root_boms.pop(item_root, None)

        return item_root_boms

    def checkStock(self, product_name, qty):
        with pyodbc.connect(self.ConnectionString) as conn:  # Opens the connection
            with conn.cursor() as cursor:  # Opens the cursor
                # Summing all relevant records
                cursor.execute(
                    "SELECT SUM(QuantityOrdered) FROM ProductionOrders "
                    "WHERE OrderStatus NOT IN ('4', '6') AND Item = ?", (product_name,)
                )

                result = cursor.fetchone()[0]  # Fetch the sum result
                # If the result is None (no records found), set stock to 0
                if result is not None:
                    item_stock = result  # This will contain the sum of the values
                else:
                    item_stock = 0

        for instance in self.Stock:
            if instance.Item == product_name:
                
                if (instance.StockAvailable + item_stock) > 0:
                    if (instance.StockAvailable + item_stock) >= qty:
                        instance.StockAvailable -= qty
                        return 0
                    else:
                        qty -= instance.StockAvailable
                        instance.StockAvailable = 0
        return qty

    def createExecutionPlans(self, main_item, prod_order):
        no_routings = []
        no_boms = []
        
        machines_Torc = [machine.MachineCode for machine in self.TorcMachines if machine.IsActive]
        machines_Tref = [machine.MachineCode for machine in self.TrefMachines if machine.IsActive]
        
        def has_routing(item_name, machines):
            """Check if there is a valid routing for the given item and machine list."""
            return any(routing.Item == item_name and routing.Machine in machines for routing in self.Routings)

        def has_bom(item_name):
            """Check if there is a valid BoM for the given item."""
            return any(bom.ItemRoot == item_name for bom in self.BoMs)
        
        def create_execution_plan(parent_item, item, quantity, bom_id, prod_order):
            """Create and add execution plan."""
            ep = ExecutionPlan(parent_item, item, float(quantity), bom_id, prod_order)
            self.ExecutionPlans.append(ep)
        
        def process_bom_items(bom, quantity, prod_order, is_main_item=False, parent_item=None):
            main_item = Items.get_Item(bom.ItemRoot, self.Database)
            
            # Create execution plan for main item if it's the first iteration
            if is_main_item and has_routing(main_item.Name, machines_Torc):
                create_execution_plan(None, main_item, quantity, None, prod_order)
            elif is_main_item and not has_routing(main_item.Name, machines_Torc):
                no_routings.append(main_item.Name)
                return

            # Process each BpM item
            for BoM_Item in bom.BoMItems:
                item = Items.get_Item(BoM_Item.ItemRelated, self.Database)
                production_qty = (BoM_Item.NetQuantity * quantity) / bom.BoMQuantity

                # Determine machine list based on parent item (tref or torc)
                machine_list = machines_Torc if parent_item is None or parent_item.Process == "BUN" else machines_Tref

                if item.Process == "BUN":
                    # Process each quantity
                    for _ in range(BoM_Item.Quantity):
                        if not has_routing(item.Name, machine_list):
                            no_routings.append(item.Name)
                            continue
                            
                        if not has_bom(item.Name):
                            no_boms.append(item.Name)
                            continue
                            
                        # Create execution plan for the current item
                        parent_for_ep = main_item if parent_item is None else parent_item
                        create_execution_plan(parent_for_ep, item, production_qty, bom.id, prod_order)
                        
                        # Recursively process sub-BoMs
                        for sub_bom in self.BoMs:
                            if sub_bom.ItemRoot == item.Name:
                                create_execution_plan(sub_bom, production_qty, prod_order, False, item)
                else:
                    # Create execution plans for items that are not BUN
                    for _ in range(BoM_Item.Quantity):
                        if not has_routing(item.Name, machines_Tref):
                            no_routings.append(item.Name)
                            continue
                            
                        if not has_bom(item.Name):
                            no_boms.append(item.Name)
                            continue
                            
                        # Create execution plan for the current item
                        parent_for_ep = main_item if parent_item is None else parent_item
                        create_execution_plan(parent_for_ep, item, production_qty, bom.id, prod_order)

        # Check if main item has BoM
        if not has_bom(main_item.Name):
            no_boms.append(main_item.Name)
            return no_routings, no_boms

        # Calculate quantities
        order_increment = main_item.OrderIncrement
        total_qty, remainder_qty = divmod(prod_order.Quantity, order_increment)

        # Process quantities
        for _ in range(int(total_qty)):
            first_iteration = True
            for bom in self.BoMs:
                if bom.ItemRoot == main_item.Name:
                    process_bom_items(bom, order_increment, prod_order, first_iteration)
                    if first_iteration:
                        first_iteration = False

        if remainder_qty > 0:
            first_iteration = True
            for bom in self.BoMs:
                if bom.ItemRoot == main_item.Name:
                    process_bom_items(bom, remainder_qty, prod_order, first_iteration)
                    if first_iteration:
                        first_iteration = False

        return no_routings, no_boms
    
    def createRemainingExecutionPlans(self, PT_Settings):
        Tref_items = {}

        for TU in self.TimeUnits:
            for exec_plan in TU.ExecutionPlans:
                item_name = exec_plan.ItemRelated.Name
                Tref_items[item_name] = round(Tref_items.get(item_name, 0) + exec_plan.Quantity)

        for tref_name in Tref_items:
            tref_item = Items.get_Item(tref_name, self.Database)
            for bom in self.BoMs:
                if bom.ItemRoot == tref_item.Name:
                    for BoM_Item in bom.BoMItems:
                        ROD_item = Items.get_Item(BoM_Item.ItemRelated, self.Database)
                        for _ in range(tref_item.Input):
                            prod_qty = int(self.checkStock(tref_item.Name, ROD_item.OrderIncrement)
                                             if self.Criteria[3] else ROD_item.OrderIncrement)
                            if PT_Settings:
                                prod_order = next((exec_plan.ProductionOrder for exec_plan in self.ExecutionPlans
                                                   if tref_item.Name == exec_plan.ItemRelated.Name))
                                if prod_qty != 0:      
                                    ep = ExecutionPlan(tref_item, ROD_item, ROD_item.OrderIncrement, bom.id,
                                                  prod_order)
                                    self.ExecutionPlans.append(ep)
        
        # After creating all execution plans, remove all instances where Process isnt ROD, MDW or BUN   
        plans_to_exclude = []     
        for exec_plan in self.ExecutionPlans:
            if exec_plan.ItemRelated.Process is None or exec_plan.ItemRelated.Process not in ["ROD", "MDW", "BUN"]:
                plans_to_exclude.append(exec_plan.id)
                
        for ep_id in plans_to_exclude:
            self.removeEPbyID(ep_id) 

    def readDBData(connection_string, database):
        def fetch_data(queries):
            data = {}
            try:
                with pyodbc.connect(connection_string) as connection:
                    with connection.cursor() as cursor:
                        for query_key, query in queries.items():
                            cursor.execute(query)
                            rows = cursor.fetchall()
                            data[query_key] = rows
            except pyodbc.Error as ex:
                print(f'Error: {ex}')
            
            return data

        def process_bom_data(rows, default):
            bom_dict = {}

            for row in rows:
                main_item = row[0]
                item = row[4] if default else row[5]
                quantity = row[1] if default else row[2]
                quantity_unit = row[2] if default else row[3]
                net_quantity = row[5] if default else row[6]
                net_quantity_unit = row[6] if default else row[7]
                revision = None if default else row[1]

                if main_item not in bom_dict:
                    bom_dict[main_item] = {
                        "BoMQuantity": quantity,
                        "BoMQuantityUnit": quantity_unit,
                        "items": {} if default else {},
                        "revisions": {} if not default else None,
                    }

                container = bom_dict[main_item]["items"] if default else bom_dict[main_item]["revisions"]
                if not default:
                    if revision not in container:
                        container[revision] = {}
                    container = container[revision]

                if item not in container:
                    container[item] = {
                        "NetQuantity": net_quantity,
                        "NetQuantityUnit": net_quantity_unit,
                        "count": 0
                    }

                container[item]["count"] += 1

            return bom_dict

        def create_bom_objects(bom_dict, default):
            for main_item, bom_data in bom_dict.items():
                if default:
                    # Create a single BoM object for the default case
                    bom = BoM(
                        main_item,
                        bom_data["BoMQuantity"],
                        bom_data["BoMQuantityUnit"],
                        None,
                        default,
                    )
                    bom_items = [
                        BoMItem(
                            main_item,
                            item,
                            item_data["count"],
                            item_data["NetQuantity"],
                            item_data["NetQuantityUnit"],
                            database
                        )
                        for item, item_data in bom_data["items"].items()
                    ]
                    
                    if database == "COFACTORY_GR":
                        BoM.GR_instances.append(bom)
                    elif database == "COFACTORY_PT":
                        BoM.PT_instances.append(bom)
                    
                    bom.add_BoM_items(*bom_items)

                else:
                    # Create a separate BoM object for each revision in the non-default case
                    for revision, items in bom_data["revisions"].items():
                        bom = BoM(
                            main_item,
                            bom_data["BoMQuantity"],
                            bom_data["BoMQuantityUnit"],
                            default,
                            revision,
                        )
                        bom_items = [
                            BoMItem(
                                main_item,
                                item,
                                item_data["count"],
                                item_data["NetQuantity"],
                                item_data["NetQuantityUnit"],
                                database
                            )
                            for item, item_data in items.items()
                        ]
                        
                        if database == "COFACTORY_GR":
                            BoM.GR_instances.append(bom)
                        elif database == "COFACTORY_PT":
                            BoM.PT_instances.append(bom)
                            
                        bom.add_BoM_items(*bom_items)

        def create_routing_objects(rows):
            for row in rows:
                main_item, machine, cycle_time, weight = row
                routing = Routings(main_item, machine, cycle_time, int(weight))
                
                if database == "COFACTORY_GR":
                    Routings.GR_instances.append(routing)
                elif database == "COFACTORY_PT":
                    Routings.PT_instances.append(routing)

        def create_items_objects(rows):
            for row in rows:
                main_item, input, diameter, unit, order_increment, process, material_type = row
                process = process if isinstance(process, str) and process != "-" else None
                item = Items(main_item, material_type, unit, input, round(diameter, 3), process, order_increment)
                
                if database == "COFACTORY_GR":
                    Items.GR_instances.append(item)
                elif database == "COFACTORY_PT":
                    Items.PT_instances.append(item)
                
        def create_machines_objects(rows):
            for row in rows:
                main_item, output, input, RT = row
                if database == 'COFACTORY_PT' and main_item in {'MDW002', 'ROD004'}:
                    continue
                machine = Machines(main_item, input, output, RT)
                
                if database == "COFACTORY_GR":
                    Machines.GR_instances.append(machine)
                elif database == "COFACTORY_PT":
                    Machines.PT_instances.append(machine)

        def create_setup_times_objects(rows):
            for row in rows:
                from_material, to_material, setup_time = row
                setup = SetupTimesByMaterial(from_material, to_material, setup_time)
                
                if database == "COFACTORY_GR":
                    SetupTimesByMaterial.GR_instances.append(setup)
                elif database == "COFACTORY_PT":
                    SetupTimesByMaterial.PT_instances.append(setup)

        def create_stock_objects(rows):
            for row in rows:
                warehouse, item, stock_available, stock_allocated, stock_economic = row
                stock =Stock(warehouse, item, stock_available, stock_allocated, stock_economic)
                
                if database == "COFACTORY_GR":
                    Stock.GR_instances.append(stock)
                elif database == "COFACTORY_PT":
                    Stock.PT_instances.append(stock)

        def create_timeunits_objects(rows):
            for row in rows:
                time_unit_id, machine_code, st, cot = row
                tu_instance = TimeUnit(machine_code)
                tu_instance.id = time_unit_id
                tu_instance.ST = st
                tu_instance.CoT = cot
                
                if database == "COFACTORY_GR":
                    TimeUnit.GR_instances.append(tu_instance)
                elif database == "COFACTORY_PT":
                    TimeUnit.PT_instances.append(tu_instance)

        def create_executionplans_objects(rows):
            for row in rows:
                exec_plan_id, main_item_name, item_name, quantity, machine_code, prod_order_id, time_unit_position, st, cot, plano_id = row

                item_root = Items.get_Item(main_item_name, database) if main_item_name else None
                item_related = Items.get_Item(item_name, database)

                exec_plan = ExecutionPlan(item_root, item_related, quantity, None, prod_order_id)
                exec_plan.id = exec_plan_id
                exec_plan.Machine = machine_code
                exec_plan.Position = time_unit_position
                exec_plan.ST = st
                exec_plan.CoT = cot
                exec_plan.PlanoId = plano_id
                
                if database == "COFACTORY_GR":
                    ExecutionPlan.GR_instances.append(exec_plan)
                elif database == "COFACTORY_PT":
                    ExecutionPlan.PT_instances.append(exec_plan)

        def create_timeunit_executionplans_objects(rows):
            instances = {
                "COFACTORY_GR": (TimeUnit.GR_instances, ExecutionPlan.GR_instances),
                "COFACTORY_PT": (TimeUnit.PT_instances, ExecutionPlan.PT_instances),
            }

            time_units, execution_plans = instances.get(database, ([], []))

            for time_unit_id, execution_plan_id in rows:
                time_unit_instance = next((tu for tu in time_units if tu.id == time_unit_id), None)
                exec_plan_instance = next((ep for ep in execution_plans if ep.id == execution_plan_id), None)

                if exec_plan_instance and time_unit_instance:
                    time_unit_instance.ExecutionPlans.append(exec_plan_instance)
                    
        def create_ln_production_orders_objects(rows):
            instance_lists = {
                "COFACTORY_GR": LN_ProductionOrders.GR_instances,
                "COFACTORY_PT": LN_ProductionOrders.PT_instances
            }

            for row in rows:
                id, item_name, machine_code, quantity, st, cot = row
                item = Items.get_Item(item_name, database)

                # Check if item meets criteria
                is_valid = (
                    (item.Process in ['ROD', 'MDW'] and item.Name.startswith('D')) or
                    (item.Process == 'BUN' and item.Name.startswith('B'))
                )

                if is_valid and database in instance_lists:
                    ln_po = LN_ProductionOrders(id, item, machine_code, quantity, st, cot)
                    instance_lists[database].append(ln_po)

        queries = {
            "boms": "SELECT Boms.MainItem, Boms.Quantity, Boms.QuantityUnit, Boms.Position, Boms.Item, Boms.NetQuantity, Boms.NetQuantityUnit FROM Boms JOIN Items i ON MainItem = i.Item WHERE i.Process IN ('ROD', 'MDW', 'BUN')",
            "eboms": "SELECT Eboms.MainItem, Eboms.Revision, Eboms.Quantity, Eboms.QuantityUnit, Eboms.Position, Eboms.Item, Eboms.NetQuantity, Eboms.NetQuantityUnit FROM Eboms JOIN Items i ON MainItem = i.Item WHERE i.Process IN ('ROD', 'MDW', 'BUN')",
            "routings": "SELECT MainItem, RoutingCode, CycleTime, Priority FROM Routings JOIN Items i ON MainItem = i.Item WHERE i.Process IN ('ROD', 'MDW', 'BUN') AND (MainItem LIKE 'B%' OR MainItem LIKE 'D%')",
            "machines": "SELECT MachineCode, WindersCount, PayoffsCount, RunningTimeFactor FROM Machines WHERE (MachineCode LIKE 'BUN0%' OR MachineCode LIKE 'BMC%' OR MachineCode LIKE 'MDW0%' OR MachineCode LIKE 'ROD0%')",
            "items": "SELECT Item, StrandsNumber, StrandsDiameter, Unit, OrderIncrement, Process, MaterialType FROM Items",
            "setup_times": "SELECT FromMaterial, ToMaterial, SetupTime FROM SetupTimesByMaterial",
            "stock": "SELECT Warehouse, Item, StockAvailable, StockAllocated, StockEconomic FROM Stock",
            "timeunits": "SELECT id, Machine, StartTime, CompletionTime FROM TimeUnits",
            "execution_plans": "SELECT id, MainItem, Item, Quantity, Machine, ProductionOrder, Position, StartTime, CompletionTime, PlanoId FROM ExecutionPlans",
            "timeunit_executionplans": "SELECT tep.TimeUnitId, tep.ExecutionPlanId FROM TimeUnitExecutionPlans tep JOIN TimeUnits tu ON tep.TimeUnitId = tu.id JOIN ExecutionPlans ep ON tep.ExecutionPlanId = ep.id",
            "production_orders": "SELECT ProductionOrderNumber, Item, Routing, QuantityOrdered, ProductionStartDateTime, PlannedDeliveryDateTime FROM ProductionOrders WHERE ((Routing LIKE 'BUN0%' OR Routing LIKE 'BMC%' OR Routing LIKE 'MDW0%' OR Routing LIKE 'ROD0%') AND (OrderStatus = '4' OR OrderStatus = '6'))"
        }

        data = fetch_data(queries)

        # Process and create objects for each type of data
        bom_dict = process_bom_data(data["boms"], True)
        create_bom_objects(bom_dict, True)

        ebom_dict = process_bom_data(data["eboms"], False)
        create_bom_objects(ebom_dict, False)

        create_items_objects(data["items"])

        create_machines_objects(data["machines"])

        create_routing_objects(data["routings"])

        create_setup_times_objects(data["setup_times"])

        create_stock_objects(data["stock"])

        create_timeunits_objects(data["timeunits"])

        create_executionplans_objects(data["execution_plans"])

        create_timeunit_executionplans_objects(data["timeunit_executionplans"])
        
        create_ln_production_orders_objects(data["production_orders"])

    def writeExcelData(self, PT_Settings, detailed):
        wb = Workbook()
        ws = wb.active
    
        # Add header row
        ws.append([
            "Factory", "Routing", "Production Order ID", "Item", "Quantity",
            "Unit", "PO Start Time", "PO End Time", "Requested Delivery Date"
        ])

        grouped_plans = {}
        for exec_plan in self.ExecutionPlans:
            # Apply logic to determine the production order based on the branch and whether it is a 'special case' product
            if PT_Settings and exec_plan.ItemRoot and exec_plan.ItemRelated.Process == 'BUN':
                # Find the main execution plan related to the current execution plan's item root
                main_ep = next(
                    (ep for ep in self.ExecutionPlans if
                     exec_plan.ItemRoot.Name == ep.ItemRelated.Name and ep.ItemRoot is None),
                    None
                )
                prod_order = main_ep.ProductionOrder if main_ep else exec_plan.ProductionOrder
            else:
                prod_order = exec_plan.ProductionOrder
            
            # The detailed version contains all the calculated reels for each product and their respective quantities
            if detailed:
                routing = "122" if PT_Settings else "125"
                ws.append([
                    routing,
                    exec_plan.Machine,
                    prod_order.id,
                    exec_plan.ItemRelated.Name,
                    exec_plan.Quantity,
                    exec_plan.ItemRelated.Unit,
                    exec_plan.ST,
                    exec_plan.CoT,
                    prod_order.DD
                ])
            # The normal version has the total quantities for each product
            else: 
                # Create the key for grouping
                key = (exec_plan.Machine, exec_plan.ItemRelated.Name, prod_order.DD, exec_plan.ItemRelated.Process)

                if key not in grouped_plans:
                    grouped_plans[key] = {
                        'quantity': exec_plan.Quantity,
                        'earliest_ST': exec_plan.ST,
                        'latest_CoT': exec_plan.CoT,
                        'unit': exec_plan.ItemRelated.Unit,
                        'prod_order': prod_order
                    }
                else:
                    grouped_plans[key]['quantity'] += exec_plan.Quantity
                    grouped_plans[key]['earliest_ST'] = min(grouped_plans[key]['earliest_ST'], exec_plan.ST)
                    grouped_plans[key]['latest_CoT'] = max(grouped_plans[key]['latest_CoT'], exec_plan.CoT)
        
        if grouped_plans:
            for (machine, item_name, dd, _), plan_data in sorted(grouped_plans.items(), key=lambda x: x[0][3]):
                routing = "122" if PT_Settings else "125"
                ws.append([
                    routing,
                    machine,
                    plan_data['prod_order'].id,
                    item_name,
                    plan_data['quantity'],
                    plan_data['unit'],
                    plan_data['earliest_ST'],
                    plan_data['latest_CoT'],
                    dd
                ])
    
        # Save the workbook to a BytesIO object (in-memory) to allow its return
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    def writeDBData(self, PlanoId):
        with pyodbc.connect(self.ConnectionString) as conn:
            with conn.cursor() as cursor:
                # Insert TimeUnit instances
                insert_timeunit_query = """INSERT INTO TimeUnits (Machine, StartTime, CompletionTime) VALUES (?, ?, ?)"""
                for time_unit in self.TimeUnits:
                    time_unit.ST = time_unit.ST.replace(microsecond=0)
                    time_unit.CoT = time_unit.CoT.replace(microsecond=0)
                    cursor.execute(insert_timeunit_query, (time_unit.Machine, time_unit.ST, time_unit.CoT))

                # Insert ExecutionPlan instances
                insert_executionplan_query = """
                INSERT INTO ExecutionPlans (MainItem, Item, Quantity, Machine, ProductionOrder, Position, StartTime, CompletionTime, PlanoId)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)"""

                for exec_plan in self.ExecutionPlans:
                    exec_plan.ST = exec_plan.ST.replace(microsecond=0)
                    exec_plan.CoT = exec_plan.CoT.replace(microsecond=0)
                    main_item = exec_plan.ItemRoot.Name if exec_plan.ItemRoot else ''
                    time_unit_position = exec_plan.Position if exec_plan.Position else ''
                    exec_plan.PlanoId = PlanoId
                    cursor.execute(insert_executionplan_query, (
                        main_item, exec_plan.ItemRelated.Name, exec_plan.Quantity, exec_plan.Machine,
                        exec_plan.ProductionOrder.id, time_unit_position, exec_plan.ST, exec_plan.CoT, PlanoId
                    ))

                # Insert TimeUnitExecutionPlan instances
                insert_timeunit_execplan_query = """
                INSERT INTO TimeUnitExecutionPlans (TimeUnitId, ExecutionPlanId) VALUES (
                    (SELECT id FROM TimeUnits WHERE Machine = ? AND StartTime = ? AND CompletionTime = ?),
                    (SELECT id FROM ExecutionPlans WHERE MainItem = ? AND Item = ? AND Quantity = ? AND Machine = ? 
                    AND ProductionOrder = ? AND Position = ? AND StartTime = ? AND CompletionTime = ? AND PlanoId = ?)
                )"""
                
                for time_unit in self.TimeUnits:
                    for exec_plan in time_unit.ExecutionPlans:
                        main_item = exec_plan.ItemRoot.Name if exec_plan.ItemRoot else ''
                        time_unit_position = exec_plan.Position if exec_plan.Position else ''
                        
                        cursor.execute(
                            "SELECT id FROM TimeUnits WHERE Machine = ? AND StartTime = ? AND CompletionTime = ?",
                            (time_unit.Machine, time_unit.ST, time_unit.CoT)
                        )
                        time_unit_id = cursor.fetchone()

                        cursor.execute(
                            "SELECT id FROM ExecutionPlans WHERE MainItem = ? AND Item = ? AND Quantity = ? AND Machine = ? "
                            "AND ProductionOrder = ? AND Position = ? AND StartTime = ? AND CompletionTime = ? AND PlanoId = ?",
                            (main_item, exec_plan.ItemRelated.Name, exec_plan.Quantity, exec_plan.Machine,
                             exec_plan.ProductionOrder.id, time_unit_position, exec_plan.ST, exec_plan.CoT, PlanoId)
                        )
                        exec_plan_id = cursor.fetchone()

                        if not time_unit_id:
                            print(f"Error: No matching TimeUnit for Machine={time_unit.Machine}, StartTime={time_unit.ST}, CompletionTime={time_unit.CoT}")
                        if not exec_plan_id:
                            print(f"Error: No matching ExecutionPlan for MainItem={main_item}, Item={exec_plan.ItemRelated.Name}, "
                                  f"Quantity={exec_plan.Quantity}, Machine={exec_plan.Machine}")
                        
                        cursor.execute(insert_timeunit_execplan_query, (
                            time_unit.Machine, time_unit.ST, time_unit.CoT,
                            main_item, exec_plan.ItemRelated.Name, exec_plan.Quantity, exec_plan.Machine,
                            exec_plan.ProductionOrder.id, time_unit_position, exec_plan.ST, exec_plan.CoT, PlanoId
                        ))
    
    def setupData(self):
        """Setup the user specific data, according to the choosen database"""
        for mach in self.Machines:
            if "ROD0" in mach.MachineCode:
                self.RODMachines.append(mach)
            elif "MDW0" in mach.MachineCode:
                self.TrefMachines.append(mach)
            elif "BMC" in mach.MachineCode or "BUN0" in mach.MachineCode:
                self.TorcMachines.append(mach)

        for item in self.Items:
            if item.Process == "ROD":
                self.RODItems.append(item)
            elif item.Process == "MDW":
                self.TrefItems.append(item)
            elif item.Process == "BUN":
                self.TorcItems.append(item)

    def clearNewDataInstances(self):
        """Clear the new data instances everytime before running the algoritm"""
        self.ExecutionPlans.clear()
        self.TimeUnits.clear()
        self.ProductionOrders.clear()