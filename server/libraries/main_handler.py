import time as tm
from openpyxl import load_workbook
import time as tm
from .abort_utils import check_abort, AbortedException
from .utils import (ProductionOrder, Items, ExecutionPlan)
from .algorithms import (RODPandS, TrefPandS, TorcPandS)

def processExtrusionInput(dataHandler, file_name):
    no_routings, no_bom = [], []
    Extrusion_Input = load_workbook(file_name)
    Extrusion_Input_Active = Extrusion_Input.active
    for row in Extrusion_Input_Active.iter_rows(min_row=2, values_only=True):
        product_name, qty, due_date, weight = row[:4]
        item = Items.get_Item(product_name, dataHandler.Database)

        if not item:
            no_bom.append(product_name)
            continue

        prod_qty = int(dataHandler.checkStock(product_name, qty) if dataHandler.Criteria[3] else qty)
        
        if prod_qty == 0:
            continue

        prod_order = ProductionOrder(item, prod_qty, due_date, weight)
        dataHandler.ProductionOrders.append(prod_order)

        no_routings_items, no_bom_items = dataHandler.createExecutionPlans(item, prod_order)

        # Only add items that don't already exist
        for item_routing in no_routings_items:
            if item_routing not in no_routings:
                no_routings.append(item_routing)
                
        for item_bom in no_bom_items:
            if item_bom not in no_bom:
                no_bom.append(item_bom)
            
    if dataHandler.Criteria[5]:
        print(dataHandler.Criteria)
        plans_to_exclude = []

        # Iterate over each root item and its corresponding deactivated BoMs
        for item_root, BoMs_to_deactivate in dataHandler.Criteria[5].items():
            stop_flag = False
            plans_by_bom_id = {}
            for exec_plan in dataHandler.ExecutionPlans:
                if exec_plan.ItemRoot and exec_plan.ItemRoot.Name == item_root:
                    # Organize the execution plans by BoM ID for this root item
                    stop_flag = True
                    plans_by_bom_id.setdefault(exec_plan.BoMId, []).append(exec_plan)
                elif exec_plan.ItemRelated.Name == item_root and stop_flag:
                    break
            
            # Now process each BoM for this root item
            for _, exec_plans in plans_by_bom_id.items():
                item_list = []
                for ep in exec_plans:
                    item_list.append(ep.ItemRelated.Name)
                if any(item_list == BoMs for BoMs in BoMs_to_deactivate):
                    plans_to_exclude.extend(exec_plan.BoMId for exec_plan in exec_plans)
                # Check if all related Items with the same BoMid belong to one of the deactivated BoMs
                #if all(any(exec_plan.ItemRelated.Name in BoMs for BoMs in BoMs_to_deactivate) for exec_plan in
                #       exec_plans):
                    # Add all the execution plans belonging to the deactivated BoM to plans_to_exclude
                #    plans_to_exclude.extend(exec_plan.id for exec_plan in exec_plans)

        # Remove the execution plans by their id
        for bomId in plans_to_exclude:
            dataHandler.removeEPbyBoMID(bomId)
    
    return no_routings, no_bom

def executePandS(dataHandler, PT_Settings, user_id):
    """Execute Planning and Scheduling with abort functionality."""
    try:
        check_abort(user_id)
        
        print("A calcular...\n")
        st = tm.time()

        # Step 1: Tref Planning
        check_abort(user_id)
        st_Tref = tm.time()
        Tref = TrefPandS(DataHandler=dataHandler, user_id=user_id)
        Tref.Planning()  # Ensure Tref.Planning() checks for abort periodically
        execution_time_ROD = 0

        if PT_Settings:
            # Step 2: ROD Planning and Scheduling
            check_abort(user_id)
            dataHandler.createRemainingExecutionPlans(PT_Settings)
            st_ROD = tm.time()
            print("ROD scheduling and planning...")
            ROD = RODPandS(DataHandler=dataHandler, user_id=user_id)
            ROD.Planning()  # Ensure ROD.Planning() checks for abort periodically
            check_abort(user_id)
            ROD.Scheduling()  # Ensure ROD.Scheduling() checks for abort periodically
            et_ROD = tm.time()
            print("ROD - DONE")
            execution_time_ROD = et_ROD - st_ROD

        # Step 3: Initial Tref Scheduling
        check_abort(user_id)
        print("Initial TREF scheduling for TORC reference points...")
        Tref.Scheduling(PT_Settings)
        print("Tref - DONE")

        # Step 4: Torc Planning and Scheduling 
        check_abort(user_id)
        st_Torc = tm.time()
        late_orders = TorcPandS(DataHandler=dataHandler).LateOrders 
        et_Torc = tm.time()
        execution_time_Torc = et_Torc - st_Torc
        print("Torc - DONE")

        # Step 5: Final Tref Rescheduling
        check_abort(user_id)
        print("Final TREF scheduling based on TORC solution...")
        Tref.Scheduling(PT_Settings, True)
        et_Tref = tm.time()
        execution_time_Tref = (et_Tref - st_Tref) - execution_time_ROD

        # Finalize
        et = tm.time()
        execution_time = et - st
        print(f"Tempo de Execução - Desbastagem: {execution_time_ROD:.2f} segundos")
        print(f"Tempo de Execução - Trefilagem: {execution_time_Tref:.2f} segundos")
        print(f"Tempo de Execução - Torção: {execution_time_Torc:.2f} segundos")
        print(f"Tempo de Execução - Total: {execution_time:.2f} segundos\n")
        return late_orders
    except AbortedException as e:
        print(f"Algorithm aborted for user {user_id}: {e}")
        return []
    except Exception as e:
        print(f"Algorithm failed for user {user_id} with error: {e}")
        return []

