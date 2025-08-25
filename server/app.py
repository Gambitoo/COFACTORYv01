from flask import Flask, jsonify, request, session, send_file, send_from_directory
from werkzeug.utils import secure_filename
from flask_cors import CORS
import os
import shutil
import pandas as pd
import pyodbc
from datetime import datetime
from apscheduler.schedulers.background import BackgroundScheduler
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor
import time
from io import BytesIO
import zipfile
from openpyxl import load_workbook
from dotenv import load_dotenv
from libraries.abort_utils import (
    clear_user_abort_event, 
    set_user_abort_event,
    cleanup_user_abort_event)
from libraries.utils import (TimeUnit, ExecutionPlan, Machines, LN_ProductionOrders, DataHandler, Items)
from libraries.main_handler import executePandS, processExtrusionInput

# Load the .env file with environment variables
load_dotenv('.env')

ADMINS = os.environ.get('ADMIN', '') # Get which users have admin privileges

if ADMINS:
    ADMINS = [int(x.strip()) for x in ADMINS.split(',') if x.strip().isdigit()]
else:
    ADMINS = []

INPUT_FOLDER = os.environ.get('STORAGE_PATH')
ALLOWED_EXTENSIONS = {'xlsx'}
TEMP_CLEANUP_INTERVAL = 30  # Run the temp cleanup every 30 minutes
TEMP_FILES_LIFETIME = 3600  # Delete temp files older than 1 hour (3600 seconds)

TEMP_PLAN_SYNC = 6 # Run the plan folder sync with the DB data every 6 hours

# All existing criteria and specific user data
all_criteria, user_data = None, {}

# Instantiate the app
app = Flask(__name__)

# Set configuration from environment variables
app.config.update(
    SECRET_KEY=os.environ.get('SECRET_KEY'),
    HOST=os.environ.get('HOST', '0.0.0.0'),
    PORT=os.environ.get('PORT', '5001'),
    URL=os.environ.get('URL', 'http://localhost'),
    DRIVER=os.environ.get('DRIVER', '{ODBC Driver 17 for SQL Server}')
)

# Create thread pool and tracking dictionaries
executor = ThreadPoolExecutor(max_workers=5)
running_algorithms = {}

# Load data from both databases
connection_strings = {
    "COFACTORY_PT": f'DRIVER={app.config["DRIVER"]};SERVER={os.environ.get("COFPT_DATABASE_SERVER")};'
                    f'DATABASE={os.environ.get("COFPT_DATABASE_NAME")};UID={os.environ.get("COFPT_DATABASE_USERNAME")};'
                    f'PWD={os.environ.get("COFPT_DATABASE_PASSWORD")};'
                    f'TrustServerCertificate=yes;',
    
    "COFACTORY_GR": f'DRIVER={app.config["DRIVER"]};SERVER={os.environ.get("COFGR_DATABASE_SERVER")};'
                    f'DATABASE={os.environ.get("COFGR_DATABASE_NAME")};UID={os.environ.get("COFGR_DATABASE_USERNAME")};'
                    f'PWD={os.environ.get("COFGR_DATABASE_PASSWORD")};'
                    f'TrustServerCertificate=yes;'
}

for db_name, connection_string in connection_strings.items():
    DataHandler.readDBData(connection_string, db_name)

# Enable CORS
CORS(app, resources={r'/*': {'origins': '*'}}, supports_credentials=True)

def get_chart_data(database):
    exec_plan_instances = ExecutionPlan.GR_instances if database == "COFACTORY_GR" else ExecutionPlan.PT_instances
    time_unit_instances = TimeUnit.GR_instances if database == "COFACTORY_GR" else TimeUnit.PT_instances
    machine_instances = Machines.GR_instances if database == "COFACTORY_GR" else Machines.PT_instances
    ln_production_orders_instances = LN_ProductionOrders.GR_instances if database == "COFACTORY_GR" else LN_ProductionOrders.PT_instances

    exec_plans = [
        {
            'id': ep.id,
            'itemRoot': ep.ItemRoot.Name if ep.ItemRoot is not None else None,
            'itemRelated': ep.ItemRelated.Name,
            'quantity': ep.Quantity,
            'ST': ep.ST,
            'CoT': ep.CoT,
            'machine': ep.Machine,
            'orderIncrement': ep.ItemRelated.OrderIncrement,
        }
        for ep in exec_plan_instances
    ]

    time_units = [
        {
            'id': tu.id,
            'machine': tu.Machine,
            'ST': tu.ST,
            'CoT': tu.CoT,
            'execution_plans': [
                {
                    'id': ep.id,
                    'itemRoot': ep.ItemRoot.Name if ep.ItemRoot is not None else None,
                    'itemRelated': ep.ItemRelated.Name,
                    'quantity': ep.Quantity,
                    'ST': ep.ST,
                    'CoT': ep.CoT,
                    'machine': ep.Machine,
                    'orderIncrement': ep.ItemRelated.OrderIncrement
                }
                for ep in tu.ExecutionPlans  
            ]
        }
        for tu in time_unit_instances
    ]
    
    ln_production_orders = [
        {
            'id': ln_po.ID,
            'item': ln_po.Item.Name,
            'machine': ln_po.Routing,
            'quantity': ln_po.Quantity,
            'ST': ln_po.ST,
            'CoT': ln_po.CoT,
            'orderIncrement': ln_po.Item.OrderIncrement
        }
        for ln_po in ln_production_orders_instances
    ]

    machines = [
        {
            'name': machine.MachineCode,
            'input': machine.Input,
            'output': machine.Output,
            'RT': machine.RunningTimeFactor
        }
        for machine in machine_instances
    ]
    
    return time_units, machines, ln_production_orders

def get_new_chart_data(dataHandler, planoId = None):
    exec_plan_instances = ExecutionPlan.GR_instances if dataHandler.Database == "COFACTORY_GR" else ExecutionPlan.PT_instances
    machine_instances = Machines.GR_instances if dataHandler.Database == "COFACTORY_GR" else Machines.PT_instances
    ln_production_orders_instances = LN_ProductionOrders.GR_instances if dataHandler.Database == "COFACTORY_GR" else LN_ProductionOrders.PT_instances
    
    # Helper function to format execution plans consistently
    def format_exec_plan(ep):
        return {
            'id': ep.id,
            'itemRoot': ep.ItemRoot.Name if ep.ItemRoot is not None else None,
            'itemRelated': ep.ItemRelated.Name,
            'quantity': ep.Quantity,
            'ST': ep.ST,
            'CoT': ep.CoT,
            'machine': ep.Machine,
            'orderIncrement': ep.ItemRelated.OrderIncrement,
        }
    
    def format_plan(row, database):
        item = Items.get_Item(row[3], database)
        
        return {
            'itemRelated': item.Name,
            'quantity': row[4],
            'ST': row[6],
            'CoT': row[7],
            'machine': row[1],
            'orderIncrement': item.OrderIncrement,
        }
    
    branch_folder = os.path.join(INPUT_FOLDER, dataHandler.Database)
    
    folders = [
        folder for folder in os.listdir(branch_folder)
        if os.path.isdir(os.path.join(branch_folder, folder))
    ]
        
    for user_folder in folders:
        user_folder_path = os.path.join(branch_folder, user_folder)
        
        # Get all plan folders in the user directory with their modification times
        plan_folders = [
            folder for folder in os.listdir(user_folder_path)
            if os.path.isdir(os.path.join(user_folder_path, folder))
        ]
    
        for path_folder in plan_folders:
            if path_folder == planoId:
                detailed_plan_path = os.path.join(path_folder, "OUTPUT_MetalPlanDetailed.xlsx") 
                
                Plan = load_workbook(os.path.join(user_folder_path, detailed_plan_path))
                Plan_Active = Plan.active
                new_exec_plans = [format_plan(row, dataHandler.Database) for row in Plan_Active.iter_rows(min_row=2, values_only=True)]
    
    # Filter by planoId if it exists, otherwise use all plans
    filtered_plans = [ep for ep in exec_plan_instances if ep.PlanoId != planoId] if planoId else [ep for ep in exec_plan_instances]
    
    # Format all execution plans
    #exec_plans = [format_exec_plan(ep) for ep in filtered_plans]
    
    # Filter new plans by planoId if it exists, otherwise use all plans
    #filtered_new_plans = [ep for ep in exec_plan_instances if ep.PlanoId == planoId] if planoId else dataHandler.ExecutionPlans

    # Format plan specific execution plans
    #new_exec_plans = [format_exec_plan(ep) for ep in filtered_new_plans]

    # Format time units
    time_units = [
        {
            'id': tu.id,
            'machine': tu.Machine,
            'ST': tu.ST,
            'CoT': tu.CoT,
            'execution_plans': [
                {
                    'id': ep.id,
                    'itemRoot': ep.ItemRoot.Name if ep.ItemRoot is not None else None,
                    'itemRelated': ep.ItemRelated.Name,
                    'quantity': ep.Quantity,
                    'ST': ep.ST,
                    'CoT': ep.CoT,
                    'machine': ep.Machine,
                    'orderIncrement': ep.ItemRelated.OrderIncrement,
                }
                for ep in tu.ExecutionPlans  # Access the ExecutionPlans associated with each TimeUnit
            ]
        }
        for tu in dataHandler.TimeUnits
    ]
    
    ln_production_orders = [
        {
            'id': ln_po.ID,
            'item': ln_po.Item.Name,
            'machine': ln_po.Routing,
            'quantity': ln_po.Quantity,
            'ST': ln_po.ST,
            'CoT': ln_po.CoT,
            'orderIncrement': ln_po.Item.OrderIncrement
        }
        for ln_po in ln_production_orders_instances
    ]

    # Format machines
    machines = [
        {
            'name': machine.MachineCode,
            'input': machine.Input,
            'output': machine.Output,
            'RT': machine.RunningTimeFactor
        }
        for machine in machine_instances
    ]
    
    return new_exec_plans, time_units, machines, ln_production_orders

def get_ep_by_plano_id(database, plano_id):
    exec_plan_instances = ExecutionPlan.GR_instances if database == "COFACTORY_GR" else ExecutionPlan.PT_instances

    filtered_instances = [ep for ep in exec_plan_instances if ep.PlanoId == plano_id]

    if not filtered_instances:
        return []  # Return an empty list if no matching execution plans are found

    min_ST = min(ep.ST for ep in filtered_instances)
    max_CoT = max(ep.CoT for ep in filtered_instances)

    return [min_ST, max_CoT]

@app.route('/selectBranch', methods=['POST'])
async def select_branch():
    # Get the selected branch
    data = request.json
    selected_branch = data.get('branch')
    user_id = data.get('userId')

    global user_data
    
    if not user_id:
        return jsonify({'status': 'error', 'message': 'User ID não fornecido.'}), 400

    # Save the user ID in the session
    session['user_id'] = user_id
    
    # Update the input and temp folder paths for the specific user (session)
    branch_folder = os.path.join(INPUT_FOLDER, selected_branch)
    upload_folder = os.path.join(branch_folder, user_id)
    temp_folder = os.path.join(upload_folder, 'temp')
    
    # Store paths in session for request context
    session['branch_folder'] = branch_folder
    session['temp_folder'] = temp_folder
    session['machines_to_remove'] = []
    session['BoMs_to_remove'] = []

    # Create the folders
    os.makedirs(upload_folder, exist_ok=True)
    os.makedirs(temp_folder, exist_ok=True)
    
    if selected_branch != "COFACTORY_PT" and selected_branch != "COFACTORY_GR":
        return jsonify({'status': 'error', 'message': 'Unidade de produção inválida.'}), 400
    
    # Get the connection string, which contains the server, chosen database, and the username and password to access it
    connection_string = connection_strings[selected_branch]
    
    # Store data in the global dictionary instead of session
    user_data[user_id] = {
        "input_data": DataHandler(selected_branch, connection_string),
        "branch_folder": None,
        "input_folder": None,
        "temp_folder": None
    }
    
    # Set up the data handler
    user_data[user_id]['input_data'].setupData() 
        
    return jsonify({
        'status': 'success', 
        'message': f'Base de dados {selected_branch} selecionada.'
    })

@app.route('/getChartData', methods=['GET'])
def get_all_data():
    response_object = {'status': 'success'}
    
    user_id = session.get('user_id')
    dataHandler = user_data[user_id]['input_data']
    
    # Get the necessary data to populate the Gantt Chart
    time_units, machines, ln_production_orders = get_chart_data(dataHandler.Database)
    
    # Populate the response object with the execution plans, time units, and machine data
    #response_object['exec_plans'] = exec_plans
    response_object['time_units'] = time_units
    response_object['machines'] = machines
    response_object['production_orders'] = ln_production_orders
    return jsonify(response_object)

@app.route('/getNewChartData', methods=['GET'])
def get_results_data():
    plano_id = request.args.get('planoId')
    
    user_id = session.get('user_id')
    dataHandler = user_data[user_id]['input_data']
    response_object = {'status': 'success'}
    
    # Get the newly created data (algorithm results) to populate the new Gantt Chart
    new_exec_plans, time_units, machines, ln_production_orders = get_new_chart_data(dataHandler, plano_id)
    #response_object['exec_plans'] = exec_plans
    response_object['new_exec_plans'] = new_exec_plans
    response_object['time_units'] = time_units
    response_object['machines'] = machines
    response_object['production_orders'] = ln_production_orders
    return jsonify(response_object)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def cleanup_temp_folder():
    """Delete useless temp files from the temp folder of each user."""
    now = time.time()
    
    for _, user_info in user_data.items():
        if "temp_folder" in user_info and user_info["temp_folder"]:
            for filename in os.listdir(user_info['temp_folder']):
                file_path = os.path.join(user_info['temp_folder'], filename)
                if os.path.isfile(file_path):
                    file_age = now - os.path.getmtime(file_path)
                    if file_age > TEMP_FILES_LIFETIME:
                        os.remove(file_path)
                        print(f"Deleted old temp file: {file_path}")
                        
def sync_plan_folders_with_db():
    """Sync plan folders with database records.
    Delete folders that don't have corresponding execution plans in the database."""
    
    # Check for both databases
    for db_name, connection_string in connection_strings.items():
        try:
            # Get all plan ids from the database
            conn = pyodbc.connect(connection_string)
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT PlanoId FROM ExecutionPlans WHERE PlanoId IS NOT NULL")
            plano_ids = set(row[0] for row in cursor.fetchall())
            cursor.close()
            conn.close()
            
            # Get all plan folders for this database
            branch_folder = os.path.join(INPUT_FOLDER, db_name)
            if not os.path.exists(branch_folder):
                continue
            
            folders = [
                folder for folder in os.listdir(branch_folder)
                if os.path.isdir(os.path.join(branch_folder, folder))]
                
            for user_folder in folders:
                user_folder_path = os.path.join(branch_folder, user_folder) 
                    
                plan_folders = [
                    folder for folder in os.listdir(user_folder_path)
                    if os.path.isdir(os.path.join(user_folder_path, folder)) and folder != "temp"
                ]
                
                for plan_folder in plan_folders:
                    if plan_folder not in plano_ids:
                        plan_folder_path = os.path.join(user_folder_path, plan_folder)
                        try:
                            shutil.rmtree(plan_folder_path)
                        except Exception as ex:
                            print(f"Error deleting folder {plan_folder_path}: {ex}")
            
            if "GR" in db_name:
                ExecutionPlan.GR_instances = [ep for ep in ExecutionPlan.GR_instances if ep.PlanoId in plano_ids]
            if "PT" in db_name:
                ExecutionPlan.PT_instances = [ep for ep in ExecutionPlan.PT_instances if ep.PlanoId in plano_ids]
                        
        except pyodbc.Error as ex:
            print(f"Error synchronizing plan folders for {db_name}: {ex}")
            
    print("Plan folder synchronization completed.")

def validate_file(file_path):
    # Expected column types for the input file
    EXPECTED_COLUMNS = {
        "Item": str,        
        "Quantity": int,       
        "StartDate": "datetime64[ns]",        
        "Priority": int,   
    }
    
    try:
        df = pd.read_excel(file_path, engine="openpyxl")

        # Check if all required data columns exist        
        if set(df.columns) != set(EXPECTED_COLUMNS.keys()):
            return False
        
        df = df.infer_objects() 
        
        # Check if each column has the correct data type
        for column, expected_type in EXPECTED_COLUMNS.items():
            if expected_type == int:
                if not pd.api.types.is_integer_dtype(df[column]):
                    return False
            elif expected_type == str:
                if not pd.api.types.is_string_dtype(df[column]):
                    return False
            elif expected_type == "datetime64[ns]":
                if not pd.api.types.is_datetime64_any_dtype(df[column]):
                    return False

        return True
    except Exception as e:
        print(e)
        return False
    
@app.route('/uploadInputFile', methods=['POST'])
def upload_file():    
    file = request.files['file']
    user_id = session.get('user_id', None)
    
    # Validate if a file was selected
    if file.filename == '':
        return jsonify({'message': 'Nenhum ficheiro selecionado.'}), 400
    
    # Validate the file format and process the file if valid
    if file and allowed_file(file.filename):
        # Secure the filename, cleaning any special chars
        plan_file = secure_filename(file.filename)
        
        # Store the file temporarily
        plan_file_path = os.path.join(session['temp_folder'], plan_file) 
        file.save(plan_file_path)
        
        # Validate the input file data
        is_valid = validate_file(plan_file_path)
        if not is_valid:
            os.remove(plan_file_path)
            return jsonify({'message': 'Ficheiro inválido, tente outra vez.'}), 400
        
        session['input_file'] = plan_file_path
        dataHandler = user_data[user_id]['input_data']
        dataHandler.CurrentTime = None
        
        # Return success response with the file path
        return jsonify({'message': 'Ficheiro lido e guardado com sucesso.', 'file_path': plan_file_path}), 200

    # Return error response if the file format is invalid
    return jsonify({'message': 'O formato do ficheiro é inválido.'}), 400

@app.route('/deleteInputFile', methods=['POST'])
def delete_file():     
    user_id = session.get('user_id', None)
    input_file = session.get('input_file', None)
    
    if user_id and input_file:
        os.remove(session['input_file'])
        
    return jsonify({'status': 'success'}), 200

@app.route('/criteria', methods=['POST'])
def process_criteria():
    global all_criteria
    data = request.json
    selected_criteria = data.get('selectedCriteria')
    criteria = data.get('allCriteria')
    
    user_id = session.get('user_id', None)
    dataHandler = user_data[user_id]['input_data']    
    all_criteria = criteria
    
    # Parse the criteria, from string to int
    parsed_criteria = {int(k): v for k, v in selected_criteria.items()}

    # Save the criteria in the dataHandler of the respective user
    dataHandler.Criteria = parsed_criteria
    print(parsed_criteria)
    print(dataHandler.Criteria)
    
    # Add the criteria to the temp folder
    with open(os.path.join(session['temp_folder'], "criteria.txt"), "w") as f:
        selected_criteria = []

        for idx, criteria in dataHandler.Criteria.items():
            if criteria:
                # Add the basic criterion name
                text = all_criteria[idx]
                print(text)

                # For removing machines or BoMs, add details
                if idx == 0 or idx == 5:
                    text += f": {criteria}"

                selected_criteria.append(text)

        # Join all criteria with commas
        f.write(", \n".join(selected_criteria))

    return jsonify({'status': 'success', 'criteria': selected_criteria}), 200

@app.route('/machines', methods=['GET'])
def get_machines():
    user_id = session.get('user_id', None)
    dataHandler = user_data[user_id]['input_data']
    machine_instances = Machines.GR_instances if dataHandler.Database == "COFACTORY_GR" else Machines.PT_instances
    
    machines = [
        {
            'name': machine.MachineCode,
            'input': machine.Input,
            'output': machine.Output,
            'RT': machine.RunningTimeFactor
        }
        for machine in machine_instances
    ]
    
    processes = ['ROD', 'MDW', 'BUN']
            
    return jsonify({'status': 'success', 'machines': machines, 'processes': processes})

@app.route('/BoMs', methods=['GET'])
def get_BoMs():
    user_id = session.get('user_id', None)
    dataHandler = user_data[user_id]['input_data']
    
    # Get the BoM's from the products in the inputted production orders
    item_BoMs = dataHandler.getInputBoMs(session['input_file'])

    return jsonify({'status': 'success', 'item_BoMs': item_BoMs})

@app.route('/removeMachines', methods=['POST'])
def remove_machines():
    machines_to_remove = request.json
    user_id = session.get('user_id', None)
    dataHandler = user_data[user_id]['input_data']
    
    # Eensure all previous deactivated machines are reactivated
    for machine in dataHandler.Machines:
        if not machine.IsActive:
            machine.IsActive = True

    # Remove the chosen machines, by deactivating them
    for machine_name in machines_to_remove:
        machine = next((m for m in dataHandler.Machines if machine_name == m.MachineCode), None)
        if machine:
            machine.IsActive = False
            
    # Add to the user's session the machines that should be deactivated    
    session['machines_to_remove'] = machines_to_remove
    
    return jsonify({'status': 'success'}), 200 # A success message is returned, if the machines are removed successfully.

@app.route('/removeBoMs', methods=['POST'])
def remove_boms():
    boms_to_remove = request.json
    
    # Add to the user's session the BoM's that should disregarded
    session['BoMs_to_remove'] = boms_to_remove

    return jsonify({'status': 'success'}), 200 # A success message is returned, if the BoM's are removed successfully.

@app.route('/createData', methods=['POST'])
def create_data():
    user_id = session.get('user_id', None)
    dataHandler = user_data[user_id]['input_data']
    
    if dataHandler.CurrentTime is None:
        dataHandler.CurrentTime = datetime.now()
        
    # Add to the algorithm criteria the machines and BoM's that should disregarded
    if 'machines_to_remove' in session:
        dataHandler.Criteria[0] = session['machines_to_remove']

    if 'BoMs_to_remove' in session:
        dataHandler.Criteria[5] = session['BoMs_to_remove'] 
    
    # Clear the user-specific data instances
    dataHandler.clearNewDataInstances()
    
    # Process input file
    no_routings, no_bom = processExtrusionInput(dataHandler, session['input_file'])
    
    return jsonify({
        'status': 'success',
        'no_routings': no_routings,
        'no_bom': no_bom,
    }), 200
    
def run_algorithm_in_thread(user_id, dataHandler, PT_Settings):
    """Run the algorithm in a separate thread"""
    try:
        # Clear any abort flag
        clear_user_abort_event(user_id)
        
        # Run the algorithm
        late_orders = executePandS(dataHandler, PT_Settings, user_id)
        
        # Update the algorithm status
        if running_algorithms[user_id]["status"] != "aborted":
            running_algorithms[user_id]["status"] = "completed"
            running_algorithms[user_id]["success"] = True
            running_algorithms[user_id]["message"] = "Algoritmo executado com sucesso." 
            running_algorithms[user_id]["late_orders"] = late_orders
            saveData(user_id, dataHandler)
    except Exception as e:
        import traceback
        # Update status with error information
        running_algorithms[user_id]["status"] = "error"
        running_algorithms[user_id]["success"] = False
        running_algorithms[user_id]["message"] = str(e)
        running_algorithms[user_id]["traceback"] = traceback.format_exc()
    finally:
        cleanup_user_abort_event(user_id)
        
        # Keep the result for a while so the client can query it
        def cleanup():
            time.sleep(300)  # Keep results for 5 minutes
            if user_id in running_algorithms:
                del running_algorithms[user_id]
                
        threading.Thread(target=cleanup).start()

@app.route('/runAlgorithm', methods=['POST'])
def run_algorithm():
    # Get the user ID from the session
    user_id = session.get('user_id') or request.args.get('user_id')
    
    # Check if the algorithm is already running for this user
    if user_id in running_algorithms and running_algorithms[user_id]["status"] == "running":
        return jsonify({
            'status': 'already_running',
            'message': 'O algoritmo já está em execução para este utilizador.'
        }), 400
        
    # Check if the thread pool is full
    if executor._work_queue.qsize() >= executor._max_workers:
        return jsonify({
            'status': 'pool_full',
            'message': 'A execução do algoritmo está em capacidade máxima. Por favor, tente novamente mais tarde.'
        }), 503  
    
    dataHandler = user_data[user_id]['input_data']
    
    # Check if PT_Settings should be enabled
    PT_Settings = True if dataHandler.Database == "COFACTORY_PT" else False
    
    # Initialize the algorithm status
    running_algorithms[user_id] = {
        "id": str(uuid.uuid4()),
        "status": "running",
        "success": None,
        "message": "Algoritmo em execução.",
        "start_time": time.time(),
        "late_orders": []
    }
    
    user_data[user_id]['branch_folder'] = session['branch_folder']
    user_data[user_id]['input_file'] = session.get('input_file', None) 
    user_data[user_id]['temp_folder'] = session['temp_folder']

    # Submit the algorithm to the thread pool
    executor.submit(run_algorithm_in_thread, user_id, dataHandler, PT_Settings)
    #status, late_orders = executePandS(dataHandler, PT_Settings)
    
    return jsonify({
        'status': 'success',
        'message': 'Iniciada a execução do algoritmo.'
    }), 202 

@app.route('/algorithmStatus', methods=['GET'])
def algorithm_status():
    user_id = session.get('user_id') or request.args.get('user_id')
    
    if not user_id:
        return jsonify({
            'status': 'error', 
            'message': 'User ID não encontrado. Por favor atualize a página.'
        }), 400
    
    # Log for debugging
    print(f"Status check for user_id: {user_id}, Active algorithms: {list(running_algorithms.keys())}")
    
    # Check if there's a task for this user
    if user_id not in running_algorithms:
        return jsonify({'status': 'not_running', 'message': 'Este utilizador não tem nenhum algoritmo em execução.'}), 200 
    
    # Return the algorithm status
    status = running_algorithms[user_id]
    
    # Check if the algorithm has finished or crashed
    if status["status"] in ["completed", "error", "aborted"]:
        # Remove it after a short delay to allow the client to receive the result
        def delayed_removal():
            time.sleep(5)  
            if user_id in running_algorithms:
                del running_algorithms[user_id]
                print(f"Removed finished algorithm for user {user_id}")
        
        threading.Thread(target=delayed_removal).start()

    return jsonify({
        'status': status["status"],
        'late_orders': running_algorithms[user_id]["late_orders"]
    }), 202 
    
@app.route('/activeAlgorithms', methods=['GET'])
def active_algorithms():
    # Check what users currently have active algorithms
    active_users = [user_id for user_id, data in running_algorithms.items() if data["status"] == 'running']
    
    return jsonify({'active_users': active_users}), 202 

@app.route('/abortAlgorithm', methods=['POST'])
def abort_algorithm():
    # Check if this is an admin terminating another user's algorithm
    data = request.get_json() if request.is_json else {}
    target_user_id = data.get('user_id') if data else None
    current_user_id = session.get('user_id')
    
    print("TARGET", target_user_id)
    print("CURRENT", current_user_id)
    
    # If no target user specified, use current user (self-termination)
    user_id = target_user_id if target_user_id else current_user_id
    
    if not user_id:
        return jsonify({'status': 'error', 'message': 'User ID não encontrado.'}), 400
    
    # If terminating another user's algorithm, check admin privileges
    if target_user_id and target_user_id != current_user_id:
        if int(current_user_id) not in ADMINS:
            return jsonify({'status': 'error', 'message': 'Acesso negado. É necessário privilégios de administrador.'}), 403
        
    print("USER_ID", user_id)
    
    # Check if there's an algorithm running for this user
    if user_id in running_algorithms and running_algorithms[user_id]["status"] == "running":
        # Set the abort flag
        set_user_abort_event(user_id)
        
        # Update the algorithm status
        running_algorithms[user_id]["status"] = "aborted"
        running_algorithms[user_id]["success"] = False
        running_algorithms[user_id]["message"] = f"A execução do algoritmo foi abortada {'pelo administrador' if target_user_id else ''}."
        
        if user_id in user_data and 'temp_folder' in user_data[user_id] and user_data[user_id]['temp_folder']:
            for filename in os.listdir(user_data[user_id]['temp_folder']):
                file_path = os.path.join(user_data[user_id]['temp_folder'], filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)
                
        print(f"A execução do algoritmo foi abortada {'pelo administrador' if target_user_id else ''}.")
            
        return jsonify({'status': "aborted", 'message': f"O algoritmo {'do utilizador ' + user_id if target_user_id else ''} foi abortado."}), 200
    
    # If we get here, either there's no algorithm for this user or it's already completed
    return jsonify({'status': 'not_found', 'message': f"Não foi encontrado nenhum algoritmo em execução {'para o utilizador ' + user_id if target_user_id else ''}."}), 404

def saveData(user_id, dataHandler):
    # Get the minimum ST value from all execution plans
    min_ST = min(plan.ST for plan in dataHandler.ExecutionPlans)
    
    formatted_ST = int(min_ST.strftime("%d%m%Y%H%M%S"))
    
    plano_id = f"{formatted_ST}_{user_id}"
    
    # Get the user's folder
    user_folder = os.path.join(user_data[user_id]['branch_folder'], user_id)
    
    # Move file from temporary storage to permanent location
    input_file = user_data[user_id]['input_file']  
    if input_file and os.path.exists(input_file):
        new_plan_folder = os.path.join(user_folder, plano_id)
        os.makedirs(new_plan_folder, exist_ok=True)
        new_plan_path = os.path.join(new_plan_folder, f"Plano_{formatted_ST}.xlsx")
        os.rename(input_file, new_plan_path)
        
        # Also move criteria file
        criteria_temp_path = os.path.join(user_data[user_id]['temp_folder'], "criteria.txt")
        new_criteria_path = os.path.join(new_plan_folder, "criteria.txt")
        os.rename(criteria_temp_path, new_criteria_path)
    else:
        return jsonify({'status': 'not_found', "message": "Ficheiro não encontrado."}), 404    
    
    # Organize the created plan in two Excel files, and write them to the plan folder
    PT_Settings = True if dataHandler.Database == "COFACTORY_PT" else False
    PO_Excel = dataHandler.writeExcelData(PT_Settings, False)
    detailed_PO_Excel = dataHandler.writeExcelData(PT_Settings, True)
    
    # File names and their content
    file_names = {"OUTPUT_MetalPlan.xlsx": PO_Excel, "OUTPUT_MetalPlanDetailed.xlsx": detailed_PO_Excel}
    
    # Write the plan Excel files to their respective plan folder
    for filename, file_obj in file_names.items():
        file_path = os.path.join(new_plan_folder, filename)
        with open(file_path, "wb") as f:
            f.write(file_obj.getvalue())

    # Save plan in the correct DB
    dataHandler.writeDBData(plano_id)
    
    # Create a zip file with both output files
    zip_file_path = os.path.join(new_plan_folder, "OUTPUT_Plans.zip")

    # Create a ZIP archive for OUTPUT files
    with zipfile.ZipFile(zip_file_path, "w", zipfile.ZIP_DEFLATED) as zipf:
        for filename, file_obj in file_names.items():
            zipf.writestr(filename, file_obj.getvalue())
    
    if dataHandler.Database == "COFACTORY_GR":
        ExecutionPlan.GR_instances.extend(dataHandler.ExecutionPlans)
        TimeUnit.GR_instances.extend(dataHandler.TimeUnits)
    else: 
        ExecutionPlan.PT_instances.extend(dataHandler.ExecutionPlans)
        TimeUnit.PT_instances.extend(dataHandler.TimeUnits)

@app.route('/savePlan', methods=['POST'])
def save_plan():
    user_id = session.get('user_id', None)    
    dataHandler = user_data[user_id]['input_data']
    
    # Get the minimum ST value from all execution plans
    min_ST = min(plan.ST for plan in dataHandler.ExecutionPlans)
    
    formatted_ST = int(min_ST.strftime("%d%m%Y%H%M%S"))
    
    plano_id = f"{formatted_ST}_{user_id}"
    
    # Get the user's folder
    user_folder = os.path.join(session['branch_folder'], user_id)
    
    # Move file from temporary storage to permanent location
    input_file = session.get('input_file', None)  
    if input_file and os.path.exists(input_file):
        new_plan_folder = os.path.join(user_folder, plano_id)
        os.makedirs(new_plan_folder, exist_ok=True)
        new_plan_path = os.path.join(new_plan_folder, f"Plano_{formatted_ST}.xlsx")
        os.rename(input_file, new_plan_path)
        session['input_file'] = None
        
        # Also move criteria file
        criteria_temp_path = os.path.join(session['temp_folder'], "criteria.txt")
        new_criteria_path = os.path.join(new_plan_folder, "criteria.txt")
        os.rename(criteria_temp_path, new_criteria_path)
    else:
        return jsonify({'status': 'not_found', "message": "Ficheiro não encontrado."}), 404    
        
    # Organize the created plan in two Excel files, and write them to the plan folder
    PT_Settings = True if dataHandler.Database == "COFACTORY_PT" else False
    PO_Excel = dataHandler.writeExcelData(PT_Settings, False)
    detailed_PO_Excel = dataHandler.writeExcelData(PT_Settings, True)
    
    # File names and their content
    file_names = {"OUTPUT_MetalPlan.xlsx": PO_Excel, "OUTPUT_MetalPlanDetailed.xlsx": detailed_PO_Excel}
    
    # Write the plan Excel files to their respective plan folder
    for filename, file_obj in file_names.items():
        file_path = os.path.join(new_plan_folder, filename)
        with open(file_path, "wb") as f:
            f.write(file_obj.getvalue())
    
    dataHandler.writeDBData(plano_id)
    
    """with open(os.path.join(new_plan_folder, "criteria.txt"), "w") as f:
        selected_criteria = []

        for idx, criteria in dataHandler.Criteria.items():
            if criteria:
                # Add the basic criterion name
                text = all_criteria[idx]
                
                # For removing machines or BoMs, add details
                if idx == 0 or idx == 5:
                    text += f": {criteria}"

                selected_criteria.append(text)

        # Join all criteria with commas
        f.write(", \n".join(selected_criteria))
            
    activated_criteria = [all_criteria[idx] for idx, activated in dataHandler.Criteria.items() if activated]
    
    with open(os.path.join(plan_folder, "criteria.txt"), "w") as f:
        for idx, criteria in enumerate(activated_criteria):
            if idx == len(activated_criteria) - 1: # Check if it's the last element
                f.write(f"{criteria}") # Don't add a comma after the last element
            else:
                f.write(f"{criteria}, \n") # Add a comma after each element except the last one"""
    
    zip_file_path = os.path.join(new_plan_folder, "OUTPUT_Plans.zip")

    # Create a ZIP archive for OUTPUT files
    with zipfile.ZipFile(zip_file_path, "w", zipfile.ZIP_DEFLATED) as zipf:
        for filename, file_obj in file_names.items():
            zipf.writestr(filename, file_obj.getvalue())
            
    # Read the ZIP file into memory for download
    with open(zip_file_path, "rb") as f:
        zip_buffer = BytesIO(f.read())
    zip_buffer.seek(0)

    if dataHandler.Database == "COFACTORY_GR":
        ExecutionPlan.GR_instances.extend(dataHandler.ExecutionPlans)
        TimeUnit.GR_instances.extend(dataHandler.TimeUnits)
    else: 
        ExecutionPlan.PT_instances.extend(dataHandler.ExecutionPlans)
        TimeUnit.PT_instances.extend(dataHandler.TimeUnits)
        
    return send_file(zip_buffer, as_attachment=True, download_name='OUTPUT_Plans.zip', mimetype='application/zip')

@app.route('/getPlanHistory', methods=['GET'])
def plan_history():
    user_id = session.get('user_id', None)
    dataHandler = user_data[user_id]['input_data']
    
    plan_history_list = []  # Store the final structured response
    
    branch_folder = os.path.join(INPUT_FOLDER, dataHandler.Database)
    
    folders = [
        folder for folder in os.listdir(branch_folder)
        if os.path.isdir(os.path.join(branch_folder, folder))
    ]
        
    for user_folder in folders:
        user_folder_path = os.path.join(branch_folder, user_folder)
        
        # Get all plan folders in the user directory with their modification times
        plan_folders = [
            (folder, os.path.getmtime(os.path.join(user_folder_path, folder)))
            for folder in os.listdir(user_folder_path)
            if os.path.isdir(os.path.join(user_folder_path, folder))
        ]
        
        # Sort folders by modification time (latest first)
        plan_folders.sort(key=lambda x: x[1], reverse=True)
        
        # Plan folders that dont exist in the DB should be removed
        folders_to_remove = []
    
        for path_folder, _ in plan_folders:
            plan_folder_path = os.path.join(user_folder_path, path_folder)
        
            # Get temp folder plan
            if plan_folder_path == session['temp_folder'] and len(os.listdir(plan_folder_path)) != 0:
            
                # Prepare lists for input files, output files, and criteria content
                input_files = []
                criteria_contents = []
                URL = app.config["URL"]
                
                for file in os.listdir(plan_folder_path):
                    file_path = os.path.join(plan_folder_path, file)
                    if os.path.isfile(file_path):
                        file_url = f"{URL}/download/{path_folder}/{file}"

                        # Categorize files
                        if file.startswith("criteria"): # File with chosen criteria
                            try:
                                with open(file_path, "r", encoding="utf-8") as criteria_file:
                                    for line in criteria_file:
                                        criteria_contents.append(line.strip())
                            except UnicodeDecodeError:
                                with open(file_path, "r", encoding="latin-1") as criteria_file:  # Try fallback encoding
                                    for line in criteria_file:
                                        criteria_contents.append(line.strip())
                        else:
                            input_files.append(file_url) # Input file used to create the plan
                                         
                plan_history_list.append({
                    "user_id": user_folder,
                    "folder": path_folder,
                    "inputFiles": input_files,
                    "outputFiles": [],
                    "ST": None,
                    "CoT": None,
                    "criteria": criteria_contents,
                    "state": "temporary"
                })
            else:
                # Prepare lists for input files, output files, and criteria content
                input_files = []
                output_files = []
                criteria_contents = []
                URL = app.config["URL"]

                # Add the zipped file with the plans data to output files
                zip_file_path = os.path.join(plan_folder_path, "OUTPUT_Plans.zip")
                if os.path.exists(zip_file_path):
                    output_files.append(f"{URL}/download/{path_folder}/OUTPUT_Plans.zip")

                for file in os.listdir(plan_folder_path):
                    file_path = os.path.join(plan_folder_path, file)
                    if os.path.isfile(file_path):
                        file_url = f"{URL}/download/{path_folder}/{file}"

                        # Categorize files
                        if file.startswith("Plano"): # Input file used to create the plan
                            input_files.append(file_url)
                        elif file.startswith("criteria"): # File with chosen criteria
                            try:
                                with open(file_path, "r", encoding="utf-8") as criteria_file:
                                    for line in criteria_file:
                                        criteria_contents.append(line.strip())
                            except UnicodeDecodeError:
                                with open(file_path, "r", encoding="latin-1") as criteria_file:  # Try fallback encoding
                                    for line in criteria_file:
                                        criteria_contents.append(line.strip())

                # Get the plan's ST and CoT 
                ST_and_CoT = get_ep_by_plano_id(dataHandler.Database, path_folder)

                if not ST_and_CoT:
                    folders_to_remove.append(plan_folder_path)
                    continue
                
                # Append the structured data to the response list
                plan_history_list.append({
                    "user_id": user_folder,
                    "folder": path_folder,
                    "inputFiles": input_files,
                    "outputFiles": output_files,
                    "ST": ST_and_CoT[0],
                    "CoT": ST_and_CoT[1],
                    "criteria": criteria_contents,
                    "state": "created"
                })
            
    # Remove folders after completing all processing
    for folder_path in folders_to_remove:
        try:
            shutil.rmtree(folder_path) 
            print(f"Removed orphaned plan folder: {folder_path}")
        except Exception as e:
            print(f"Error removing folder {folder_path}: {e}")
            
    return jsonify(plan_history_list), 200

@app.route('/checkAdminPriviliges', methods=['GET'])
def check_admin_privileges():
    # Check if current user is an admin
    user_id = session.get('user_id', None)

    try:
        user_id_int = int(user_id) if user_id else None
        is_admin = user_id_int in ADMINS
        return jsonify({'isAdmin': is_admin}), 200
    except (ValueError, TypeError):
        return jsonify({'isAdmin': False}), 200
        

@app.route('/download/<folder_name>/<filename>', methods=['GET'])
def download_file(folder_name, filename):
    """Serves files from the user's plan history folder."""    
    # Get the user's upload folder
    user_id = folder_name.split('_', 1)[1]
    
    upload_folder = os.path.join(session['branch_folder'], user_id)
    
    plan_folder = os.path.join(upload_folder, folder_name)
    
    if not os.path.exists(os.path.join(plan_folder, filename)):
        return jsonify({"message": "Ficheiro não encontrado."}), 404

    return send_from_directory(plan_folder, filename, as_attachment=True)
   
if __name__ == '__main__':
    # Start the temp cleanup scheduler before running the app
    scheduler = BackgroundScheduler()
    
    # Add temp folder cleanup job
    scheduler.add_job(cleanup_temp_folder, 'interval', minutes=TEMP_CLEANUP_INTERVAL)
    
    # Add plan folder and DB sync job
    scheduler.add_job(sync_plan_folders_with_db, 'interval', hours=TEMP_PLAN_SYNC)
    scheduler.start()
    
    try:
        app.run(host=app.config["HOST"], port=app.config["PORT"], debug=False)
    except KeyboardInterrupt:
        scheduler.shutdown() # The scheduler is properly shut down if the app stops
        